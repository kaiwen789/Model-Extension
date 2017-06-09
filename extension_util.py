import os
import openpyxl
import pkg.Simulator.simulator as sim


def extend_unit(mdl,edges,ignore,simRun,simLen,outPath):

	if edges[0] in ignore: return None
	ignore.add(edges[0])

	trace = outPath+"~temp_trace_"+str(edges[0])+".txt"
	ext_model = outPath + "~temp_final_" + str(edges[0]) + ".xlsx"
	extend_model(mdl,edges,ext_model)

	model = sim.Manager(ext_model)
	model.run_simulation('ra',simRun,simLen,trace,outMode=3)




# Read in the model, return a dict indicating the row of each element
def getRow(mdl):
	res = dict()
	wb = openpyxl.load_workbook(mdl)
	ws = wb.active
	for idx,rows in enumerate(ws.iter_rows(),start=1):
		if rows[0].value is None or rows[0].value.lower()=="variable name": continue
		res[rows[0].value] = idx
	return res


# Open a new xlsx and extend the edges into the model, return the extended model
def extend_model(mdl,edges,ext_model):
	
	os.system('cp '+mdl+' '+ext_model)

	name_to_row = getRow(ext_model)
	curr_row = len(name_to_row)+2

	wb = openpyxl.load_workbook(ext_model)
	ws = wb.active
	for e in edges[1:]:
		if e[0] not in name_to_row: 
			ws.cell(row=curr_row,column=1,value=e[0])
			ws.cell(row=curr_row,column=6,value=1)
			name_to_row[e[0]] = curr_row
			curr_row += 1
		if e[1] not in name_to_row:
			ws.cell(row=curr_row,column=1,value=e[1])
			ws.cell(row=curr_row,column=6,value=1)
			name_to_row[e[1]] = curr_row
			curr_row += 1
		col = 2 if e[2]=='+' else 3
		original = ws.cell(row=name_to_row[e[1]],column=col).value
		original = original+',' if original!=None else ''
		ws.cell(row=name_to_row[e[1]],column=col,value=original+e[0])
	
	wb.save(ext_model)
