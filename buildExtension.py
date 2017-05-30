import re,os
import pickle
import argparse
import copy,sys
import openpyxl
from collections import defaultdict
from markovCluster import runMarkovCluster

def getType(ip):
	# Note that there are plenty of different element types
	# For now exceptions are all handled as other

	if ip.upper().startswith('PROTEIN'):
		return 'Protein'
	elif ip.upper().startswith('CHEMICAL'):
		return 'Chemical'
	elif ip.upper().startswith('BIOLOGICAL'):
		return 'Biological'
	else:
		return 'other'

# Build a graph based on the baseline model
# Return a set of all names and a dict of edges
def parseModel(mdl):

	names = set()
	regulators = dict()
	wb = openpyxl.load_workbook(mdl)
	ws = wb.active

	for rows in ws.iter_rows():
		if rows[0].value is None or rows[0].value=='Variable name': continue
		curr = []
		if rows[1].value is not None: curr += re.findall('[\w]+@[\w]+',rows[1].value)
		if rows[2].value is not None: curr += re.findall('[\w]+@[\w]+',rows[2].value)
		regulators[rows[0].value] = set(curr)
		names.add(rows[0].value)

	return names, regulators


# Get the name of the extension, return a string
def getName(dict_file, curr_map, info):
	
	query, ID, e_type = info[0], info[2], getType(info[4])
	if query in curr_map: return curr_map[query]
	
	wb = openpyxl.load_workbook(dict_file)
	ws = wb.active

	match = query + '@ext'
	confidence = 0.0
	# Iterate all names and find the most likely match
	for rows in ws.iter_rows():
		if rows[0].value is None or rows[1].value is None: continue
		# print(rows[0].value,rows[1].value,rows[2].value,rows[3].value)
		curr_conf = 0.0
		if ID.upper() in rows[1].value:
			curr_conf = 1
		elif query.upper().startswith(rows[0].value.upper()) or rows[0].value.upper().startswith(query.upper()):
			curr_conf = 0.8
		if curr_conf>0 and rows[2].value.startswith(e_type):
			curr_conf += 1

		if curr_conf > confidence:
			match = rows[3].value
			confidence = curr_conf
			if curr_conf==2: break

	curr_map[query] = match
	return match

# Return a list containing edges, ignoring duplicates
# Also return the information of each edge
def parseExtension(dict_file, base_model, ext_file):

	ext_edges, ext_info, curr_map = set(), dict(), dict()
	with open(ext_file) as f:
		for line in f:
			if line.startswith('regulator_name'): continue
			line = line.strip()
			s = re.split(',',line)
			name1, name2 = getName(dict_file,curr_map,s[0:7]), getName(dict_file,curr_map,s[7:14])
			pos = '+' if s[-3]=='increases' else '-'

			if (name2 in base_model and name1 in base_model[name2]) \
				or (name1, name2, pos) in ext_info:
				# print('duplicate: ', (name1, name2, pos)) 
				# Ignore duplicate for now
				continue
			ext_edges.add((name1, name2, pos))
			ext_info[(name1, name2, pos)] = s[15:]

	return ext_edges, ext_info



def input_parsing():

	parser = argparse.ArgumentParser()
	requiredNamed = parser.add_argument_group('required input arguments')
	requiredNamed.add_argument("-m","--model",help="name of the model (under dataPath)", required=True)
	requiredNamed.add_argument("-e","--extension",help="name of the extension (under dataPath)", required=True)
	parser.add_argument("-i", "--dataPath", help="relative path to the data", default="../data/")
	parser.add_argument("-o","--outPath", help="output folder for current experiment", default="exp")
	parser.add_argument("-d","--dictionary", help="name of the dictionary", default="dictionary.xlsx")
	parser.add_argument("-me","--method", help="method to group extensions", choices=["markovCluster"], default="markovCluster")
	additionalArg = parser.add_argument_group('additional argument for each method')
	additionalArg.add_argument("--extOnly", help="(markov cluster) use extension only, default as false", action="store_true")
	additionalArg.add_argument("--markovCoef", help="(markov cluster) coefficient for markov clustering, default as 2", default=2, type=int)
	args = parser.parse_args()

	assert os.path.isdir(args.dataPath), "dataPath \"" + args.dataPath + "\" is not a path"
	assert os.path.exists(args.dataPath+args.dictionary), "dictionary \"" + args.dataPath+args.dictionary + "\" does not exist"
	assert os.path.exists(args.dataPath+args.model), "model \"" + args.dataPath+args.model + "\" does not exist"
	assert os.path.exists(args.dataPath+args.extension), "extension \"" + args.dataPath+args.dictionary + "\" does not exist"

	num, orig, curr = 1, args.outPath, args.outPath
	while os.path.isdir(args.dataPath + curr):
		curr = orig + " " + str(num)
		num += 1
	os.makedirs(args.dataPath + curr)
	args.outPath = curr + '/'

	return args


def main():

	args = input_parsing()
	print(args)

	bsl_mdl = args.dataPath + args.model
	dict_file = args.dataPath + args.dictionary
	ext_file = args.dataPath + args.extension
	grouped_file = args.dataPath + args.outPath + 'grouped_ext'

	base_name, base_model = parseModel(bsl_mdl)
	ext_edges, ext_info = parseExtension(dict_file, base_model, ext_file)
	

	if args.method == "markovCluster":
		
		ip_model = dict() if args.extOnly else base_model
		coef = args.markovCoef
		res = runMarkovCluster(args.dataPath + args.outPath,ext_edges,ip_model,coef)
		pickle.dump(res, open(grouped_file,'wb'))
	

	


if __name__=='__main__':
	main()
	

	

	


